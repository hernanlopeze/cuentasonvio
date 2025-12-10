

import streamlit as st
import pandas as pd

import os
import io
from openpyxl import load_workbook
from mayorista_minorista import proceso_mayorista_minorista
from comprobantes_en_linea import proceso_comprobantes_en_linea
from csv_piva import proceso_csv_piva

def upload_csv():
	uploaded_file = st.file_uploader("Selecciona un archivo CSV", type=["csv"])
	if uploaded_file is not None:
		df = pd.read_csv(uploaded_file, sep=';')
		return df
	return None

def preprocess_df(df):
	# Convertir la columna 'Fecha de Emisión' a datetime con el formato correcto
	if "Fecha de Emisión" in df.columns:
		df["Fecha de Emisión"] = pd.to_datetime(df["Fecha de Emisión"].astype(str), format="%Y-%m-%d", errors="coerce")
	# Convertir columnas desde 'Tipo Cambio' en adelante a numérico
	tipo_cambio_idx = None
	for idx, col in enumerate(df.columns):
		if col.strip() == "Tipo Cambio":
			tipo_cambio_idx = idx
			break
	if tipo_cambio_idx is not None:
		for col in df.columns[tipo_cambio_idx:]:
			df[col] = pd.to_numeric(
				df[col].astype(str)
				.str.replace('.', '', regex=False)
				.str.replace(',', '.', regex=False)
				.str.replace('$','')
				.str.replace('"','')
				.str.strip(),
				errors='coerce'
			)
	# Completar columna Moneda con 'PES'
	if "Moneda" in df.columns:
		df["Moneda"] = "PES"
	return df

def proceso_mayorista_minorista(df):
	# Leer comprobantes B
	path_b = os.path.join("assets", "comprobantes_b.csv")
	comprobantes_b = pd.read_csv(path_b, sep=",", encoding="utf-8")
	codigos_b = set(comprobantes_b[comprobantes_b.columns[0]].astype(str).str.strip())

	# Agregar columnas al final
	df["Importe Reg Esp 1 (AFIP - Mis Comprobantes)"] = ""
	df["Importe Reg Esp 2 (AFIP - Mis Comprobantes)"] = ""
	df["Importe Reg Esp 3 (AFIP - Mis Comprobantes)"] = ""
	df["Importe Reg Esp 4 (AFIP - Mis Comprobantes)"] = ""
	# Asignar VTAMIN si Tipo de Comprobante (por código) está en comprobantes_b.csv, si no VTAMAY
	if "Tipo de Comprobante" in df.columns:
		df["Código de Concepto/Artículo"] = df["Tipo de Comprobante"].apply(lambda x: "VTAMIN" if str(x).strip() in codigos_b else "VTAMAY")
	else:
		df["Código de Concepto/Artículo"] = "VTAMAY"
	return df

def export_excel(df):
	output = io.BytesIO()
	with pd.ExcelWriter(output, engine='openpyxl') as writer:
		df.to_excel(writer, index=False, sheet_name='Comprobantes')
		ws = writer.sheets['Comprobantes']
		# Formatear fecha corta
		if "Fecha de Emisión" in df.columns:
			col_idx = df.columns.get_loc("Fecha de Emisión") + 1
			for cell in ws[ws.cell(row=1, column=col_idx).column_letter]:
				if cell.row == 1:
					continue
				cell.number_format = 'DD/MM/YYYY'
		# Formatear columnas desde 'Tipo Cambio' en adelante como número
		tipo_cambio_idx = None
		for idx, col in enumerate(df.columns):
			if col.strip() == "Tipo Cambio":
				tipo_cambio_idx = idx
				break
		if tipo_cambio_idx is not None:
			for col in df.columns[tipo_cambio_idx:]:
				col_idx = df.columns.get_loc(col) + 1
				for cell in ws[ws.cell(row=1, column=col_idx).column_letter]:
					if cell.row == 1:
						continue
					cell.number_format = '#,##0.00'
	output.seek(0)
	st.download_button(
		label="Descargar Excel",
		data=output,
		file_name="comprobantes_modificados.xlsx",
		mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
	)

def main():
		tabs = st.tabs(["Importación a Onvio", "Importación a Portal IVA"])

		with tabs[0]:
			st.title("Importación a Onvio")
			st.write("Sube tu archivo de comprobantes en formato CSV.")
			df = upload_csv()
			if df is not None:
				df = preprocess_df(df)
				# Asegurar que las columnas relevantes sean int
				for col in ["Tipo de Comprobante", "Punto de Venta", "Número Desde"]:
					if col in df.columns:
						df[col] = pd.to_numeric(df[col], errors="coerce").astype('Int64')
				st.success("Archivo cargado correctamente!")
				st.write("Vista previa de los primeros registros:")
				st.dataframe(df.head())

				opcion = st.selectbox("Selecciona el proceso a realizar", ["Mayorista/minorista", "Comprobantes en línea"], key="onvio_select")

				if opcion == "Mayorista/minorista":
					df_proc = proceso_mayorista_minorista(df)
					export_excel(df_proc)
				elif opcion == "Comprobantes en línea":
					st.subheader("Sube los archivos TXT de comprobantes en línea")
					txt_files = st.file_uploader("Selecciona uno o más archivos TXT", type=["txt"], accept_multiple_files=True, key="txt_files_onvio")
					if txt_files:
						rows = []
						for txt_file in txt_files:
							content = txt_file.read().decode("utf-8", errors="ignore")
							for line in content.splitlines():
								tipo = line[0:2] if len(line) >= 2 else ""
								pv = line[11:15] if len(line) >= 15 else ""
								cpte = line[15:23] if len(line) >= 23 else ""
								detalle = line[114:] if len(line) >= 114 else ""
								rows.append({"Tipo": tipo, "PV": pv, "Cbte": cpte, "Detalle": detalle})
						df_txt = pd.DataFrame(rows)
						# Asegurar que las columnas relevantes sean int
						for col in ["Tipo", "PV", "Cbte"]:
							if col in df_txt.columns:
								df_txt[col] = pd.to_numeric(df_txt[col], errors="coerce").astype('Int64')
						# Agregar columnas en blanco antes del cruce
						for col in [
							"Importe Reg Esp 1 (AFIP - Mis Comprobantes)",
							"Importe Reg Esp 2 (AFIP - Mis Comprobantes)",
							"Importe Reg Esp 3 (AFIP - Mis Comprobantes)",
							"Importe Reg Esp 4 (AFIP - Mis Comprobantes)",
							"Código de Concepto/Artículo",
							"Provincia IIBB"
						]:
							if col not in df.columns:
								df[col] = ""
						# Ordenar el df principal
						df = df.sort_values(["Punto de Venta", "Número Desde"])
						# Cruce exacto
						df = df.merge(df_txt[["Tipo", "PV", "Cbte", "Detalle"]],
									  left_on=["Tipo de Comprobante", "Punto de Venta", "Número Desde"],
									  right_on=["Tipo", "PV", "Cbte"],
									  how="left")
						df = df.drop_duplicates(subset=["Tipo", "PV", "Cbte"], keep="first")

						
						st.write("Comprobantes cruzados con Detalle:")
						
						st.dataframe(df)
						# Exportar a Excel con formato de fecha corta si corresponde
						output = io.BytesIO()
						with pd.ExcelWriter(output, engine='openpyxl') as writer:
							df.to_excel(writer, index=False, sheet_name='Comprobantes')
							ws = writer.sheets['Comprobantes']
							if "Fecha de Emisión" in df.columns:
								col_idx = df.columns.get_loc("Fecha de Emisión") + 1
								for cell in ws[ws.cell(row=1, column=col_idx).column_letter]:
									if cell.row == 1:
										continue
									cell.number_format = 'DD/MM/YYYY'
						output.seek(0)
						st.download_button(
							label="Descargar Excel",
							data=output,
							file_name="comprobantes_cruzados.xlsx",
							mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
						)
		with tabs[1]:
			proceso_csv_piva()

if __name__ == "__main__":
	main()
